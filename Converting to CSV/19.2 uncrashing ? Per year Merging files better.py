import pandas as pd
import os
from glob import glob
import re
from functools import reduce

# =========================
# CONFIG
# =========================
base_dir = "/Users/isabella/Library/Mobile Documents/com~apple~CloudDocs/Graduated/UCL EvidENT/DOWNLOADED DATA/CSV READABLE"

# Create new output folder inside CSV READABLE
output_dir = os.path.join(base_dir, "MERGED_OUTPUT")
os.makedirs(output_dir, exist_ok=True)

# Only include these folders
include_folders = [
    "Audiometry - Tympanometry",
    "Audiometry - Wideband Reflectance",
    "Audiometry - Acoustic Reflex",
    "Audiometry",
]

# Optional dependencies
try:
    import pyarrow
    csv_engine = "pyarrow"
except ImportError:
    csv_engine = "c"

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    openpyxl_installed = True
except ImportError:
    openpyxl_installed = False

# =========================
# Helper functions
# =========================
def extract_year(fname):
    # Match either 2007 or 2017-2020
    match = re.search(r"\d{4}(?:-\d{4})?", os.path.basename(fname))
    if match:
        return match.group(0).replace("-", "_")  # normalise to underscores
    return "UnknownYear"

def colour_excel(excel_file, colour_map):
    if not openpyxl_installed:
        return
    wb = load_workbook(excel_file)
    ws = wb.active
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == "SEQN":
            continue
        for key, colour in colour_map.items():
            if key in str(cell.value):
                fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1):
                    for c in row:
                        c.fill = fill
                break
    wb.save(excel_file)

# =========================
# Find CSV files
# =========================
all_csv_files = []
for folder in include_folders:
    folder_path = os.path.join(base_dir, folder)
    all_csv_files.extend(glob(os.path.join(folder_path, "*.csv")))

# Skip mapping files and only keep those with years
csv_files = [
    f for f in all_csv_files
    if "mapping" not in os.path.basename(f).lower()
    and re.search(r"\d{4}(?:-\d{4})?", os.path.basename(f))
]

if not csv_files:
    print("No relevant CSV files found. Check your directory or year patterns.")
    exit()

# =========================
# Group files by year
# =========================
files_by_year = {}
for f in csv_files:
    year = extract_year(f)
    files_by_year.setdefault(year, []).append(f)

# =========================
# Process each year
# =========================
colours = [
    "FFCCCC", "CCFFCC", "CCCCFF", "FFFFCC", "CCFFFF",
    "FFCCFF", "FFD9B3", "E6CCFF", "B3E6B3", "FFB3B3"
]

for year, files in files_by_year.items():
    print(f"\n======================")
    print(f"Processing year: {year} ({len(files)} files)")
    print(f"======================")

    dfs = []
    summary_log = []
    colour_map = {}
    colour_idx = 0

    for i, file in enumerate(files, start=1):
        print(f"[{i}/{len(files)}] Processing: {file}")

        try:
            df = pd.read_csv(file, engine=csv_engine)
        except Exception as e:
            print(f" - Failed to read {file}: {e}")
            summary_log.append(f"Failed: {file} ({e})")
            continue

        # Detect SEQN
        seqn_col = next((c for c in df.columns if "SEQN" in c.upper().strip()), None)
        if seqn_col is None:
            print(" - No SEQN column, skipping")
            summary_log.append(f"Skipped: {file} (no SEQN)")
            continue
        df = df.rename(columns={seqn_col: "SEQN"})

        # Drop duplicates
        if df["SEQN"].duplicated().any():
            print(f"WARNING: Duplicate SEQN in {file} — keeping first only")
            df = df.drop_duplicates(subset="SEQN", keep="first")

        # Add suffix from file name
        file_base = os.path.splitext(os.path.basename(file))[0].replace(" ", "_")
        df = df.rename(columns={c: f"{c}_{file_base}" for c in df.columns if c != "SEQN"})

        # Assign colour
        colour_map[file_base] = colours[colour_idx % len(colours)]
        colour_idx += 1

        dfs.append(df.set_index("SEQN"))
        summary_log.append(f"Included: {file} ({len(df)} rows, {df.shape[1]-1} cols)")

    # Skip if no valid files for this year
    if not dfs:
        print(f"No valid data for year {year}, skipping.")
        continue

    # Merge all on SEQN
    merged_df = reduce(lambda left, right: pd.merge(left, right, on="SEQN", how="outer"), dfs)

    print(f"Merged dataframe shape for {year}: {merged_df.shape}")

    # Save CSV and Excel
    output_csv = os.path.join(output_dir, f"{year}_MERGED.csv")
    output_excel = os.path.join(output_dir, f"{year}_MERGED.xlsx")
    merged_df.reset_index().to_csv(output_csv, index=False)
    merged_df.reset_index().to_excel(output_excel, index=False)
    print(f"✅ Saved CSV: {output_csv}")
    print(f"✅ Saved Excel: {output_excel}")

    # Apply colour coding to Excel
    colour_excel(output_excel, colour_map)

    # Save summary log
    log_file = os.path.join(output_dir, f"{year}_MERGE_LOG.txt")
    with open(log_file, "w") as f:
        for line in summary_log:
            f.write(line + "\n")
    print(f"📄 Saved summary log: {log_file}")
