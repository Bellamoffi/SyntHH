import pandas as pd
import os
from glob import glob
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def extract_year(fname):
    """Extract NHANES year string like 1999-2000 or 2007_2008 from a filename."""
    match = re.search(r"(19\d{2}[_-]\d{2}|\d{4}[_-]\d{4})", fname)
    if match:
        return match.group(1).replace("-", "_")
    else:
        return "UnknownYear"

def colour_excel(excel_file, colour_map):
    """Apply background colours to Excel columns based on their source file."""
    wb = load_workbook(excel_file)
    ws = wb.active

    # Map each column to a colour
    for col_idx, col_name in enumerate(ws[1], start=1):
        if col_name.value == "SEQN":
            continue  # leave SEQN uncoloured
        for key, colour in colour_map.items():
            if key in str(col_name.value):
                fill = PatternFill(start_color=colour, end_color=colour, fill_type="solid")
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=1):
                    for cell in row:
                        cell.fill = fill
                break

    wb.save(excel_file)

def merge_nhanes_by_year(base_dir, year):
    """Merge all CSVs for a given year into one dataframe and save as CSV + Excel with colour coding."""
    all_csv_files = glob(os.path.join(base_dir, "**", "*.csv"), recursive=True)
    csv_files = [f for f in all_csv_files if "mapping" not in os.path.basename(f).lower()]
    year_files = [f for f in csv_files if extract_year(f) == year]

    if not year_files:
        print(f"No CSV files found for year {year}.")
        return

    merged_df = None
    included_files_names = []
    all_columns_seen = set()
    summary_log = []
    colour_map = {}

    colours = [
        "FFCCCC", "CCFFCC", "CCCCFF", "FFFFCC", "CCFFFF",
        "FFCCFF", "FFD9B3", "E6CCFF", "B3E6B3", "FFB3B3"
    ]  # rotate colours

    colour_idx = 0

    for i, file in enumerate(year_files, start=1):
        print(f"\n[{i}/{len(year_files)}] Processing: {file}")
        try:
            df = pd.read_csv(file)
        except Exception as e:
            print(f"Failed to read {file}: {e}")
            summary_log.append(f"Failed to read: {file} ({e})")
            continue

        print(f" - Columns: {df.columns.tolist()[:10]}{'...' if len(df.columns) > 10 else ''}")
        print(f" - Number of rows: {len(df)}")

        # Detect SEQN column
        seqn_col = None
        for col in df.columns:
            if "SEQN" in col.upper().strip():
                seqn_col = col
                break
        if seqn_col is None:
            print(" - No SEQN column, skipping file")
            summary_log.append(f"Skipped: {file} (no SEQN)")
            continue

        df = df.rename(columns={seqn_col: "SEQN"})

        # Add suffix
        file_base = os.path.splitext(os.path.basename(file))[0].replace(" ", "_")
        included_files_names.append(file_base)
        new_columns = {}
        for col in df.columns:
            if col != "SEQN":
                col_renamed = f"{col}_{file_base}"
                new_columns[col] = col_renamed
                if col_renamed in all_columns_seen:
                    print(f"WARNING: Column {col_renamed} already exists")
                all_columns_seen.add(col_renamed)
        df = df.rename(columns=new_columns)

        # Assign a colour to this file
        colour_map[file_base] = colours[colour_idx % len(colours)]
        colour_idx += 1

        # Merge
        if merged_df is None:
            merged_df = df
            print(f" - Initialized merged dataframe with {len(df)} participants")
            summary_log.append(f"Included: {file}")
        else:
            merged_df = merged_df.merge(df, on="SEQN", how="outer")
            print(f" - Merged file; current merged participants: {len(merged_df)}")
            summary_log.append(f"Merged: {file}")

    if merged_df is not None:
        output_csv = os.path.join(base_dir, f"{year}_MERGED.csv")
        output_excel = os.path.join(base_dir, f"{year}_MERGED.xlsx")

        merged_df.to_csv(output_csv, index=False)
        merged_df.to_excel(output_excel, index=False)

        colour_excel(output_excel, colour_map)

        print(f"\nMERGE COMPLETE for {year}")
        print(f" - Files merged: {len(included_files_names)}")
        print(f" - Participants: {merged_df.shape[0]}")
        print(f" - Columns: {merged_df.shape[1]}")
        print(f" - Saved CSV: {output_csv}")
        print(f" - Saved Excel (coloured): {output_excel}")
    else:
        print(f"No files merged for {year}")

def main():
    print("=== NHANES Merger ===")
    base_dir = input("Enter the folder path containing your CSV files: ").strip()
    if not os.path.exists(base_dir):
        print("Invalid path. Please check and try again.")
        return

    # Detect years
    all_csv_files = glob(os.path.join(base_dir, "**", "*.csv"), recursive=True)
    detected_years = sorted({extract_year(f) for f in all_csv_files})
    print(f"\nDetected years: {', '.join(detected_years)}")

    choice = input("Enter year(s) to merge (comma separated) or 'all' for all years: ").strip()
    if choice.lower() == "all":
        years_to_process = detected_years
    else:
        years_to_process = [y.strip() for y in choice.split(",") if y.strip() in detected_years]

    for year in years_to_process:
        merge_nhanes_by_year(base_dir, year)

if __name__ == "__main__":
    main()
